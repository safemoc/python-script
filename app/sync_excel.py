from openpyxl import load_workbook
from pandas import DataFrame
import pandas as pd
import datetime
import getpass
import time
import os


class utils:

    @staticmethod
    def get_files_path(path: str, exclude: list = None, file_type: str = '.xlsx') -> list:
        """
        获取目录下所有xlsx文件 不包括特定的文件
        :param path: the path of File dir
        :param file_type: file_type
        :param exclude: filename.xxx
        :return: all
        """
        if exclude is None:
            exclude = []
        elif path is None:
            raise "Path Cannot Be Empty !"
        all_work_list = []
        for root, dirs, files in os.walk(path):
            for file in files:
                if file.endswith(file_type) and file not in exclude:
                    all_work_list += [os.path.join(root, file)]
        return all_work_list

    @staticmethod
    def get_file_name(file_path) -> str:
        """
        :param file_path: 文件路径
        :return: filename
        """
        return file_path.split('\\')[-1][:-5]

    def create_sheet_dict(self, files_list: list) -> dict:
        sheet_dict = {}
        for i in files_list:
            sheet_dict[self.get_file_name(i)] = i
        return sheet_dict

    @staticmethod
    def sheet_of_notin_file(sheet_names, file_path):
        requestions = []
        collect = load_workbook(file_path)
        for i in sheet_names:
            if i not in collect.sheetnames:
                requestions += [i]
        return requestions

    @staticmethod
    def create_sheets(sheet_name: str, file_path: str, ) -> (bool, str):
        try:
            book = load_workbook(file_path)
            book.create_sheet(sheet_name)
            book.save(file_path)
            book.close()
            return True, None
        except Exception as e:
            return False, f'{e}'

    @staticmethod
    def get_template(file_path: str, header: int = 0):
        return list(pd.read_excel(file_path, header=header, ).columns)

    @staticmethod
    def insert_sheet_template(template, sheet_name, file_path):
        book = load_workbook(file_path)
        sht = book[sheet_name]
        sht.append(template)
        book.save(file_path)
        book.close()

    @staticmethod
    def format_date_today(sheet_data: DataFrame) -> DataFrame:
        sheet_data['create_date_days'] = sheet_data['create_date'].apply(
            lambda x: datetime.datetime.strptime(str(x).strip().split(' ')[0], '%Y-%m-%d').date())  # 确保数据正常，将时间格定在天
        return data

    @staticmethod
    def set_col(origin_data: DataFrame) -> DataFrame:
        df1 = origin_data.copy()
        df1.columns = [f'o_{i}' for i in origin_data.columns]
        return df1


class syncExcel(utils):
    user = getpass.getuser()

    def __init__(self, path, exclude: str | list = None):
        """
        :param path: 文件存放目录
        :param exclude: 文件目录下的不需要同步的文件
        """
        if isinstance(exclude, str):
            exclude = [exclude]
        self.sheet_dict = self.create_sheet_dict(self.get_files_path(path, exclude))

    def create_not_exists_sheet_on_collect_file(self, collect_file_info: tuple, sheet_names: list) -> None:
        sheet_names = self.sheet_of_notin_file(sheet_names, collect_file_info[1])
        for name in sheet_names:
            self.create_sheets(name, collect_file_info[1])
            self.insert_sheet_template(self.get_template(self.sheet_dict[name]), name, collect_file_info[1])
        return None

    def get_updated_data_body(self, collect_info):

        def _(store, head, column):
            new_data = pd.merge(store, head, left_on='create_date_days', right_on='o_create_date_days',
                                how='left')
            return new_data[new_data['o_id'].isnull()].loc[:, column]

        for sheet_name, store_file_path in self.sheet_dict.items():
            collect_data = pd.read_excel(collect_info[1], header=0, sheet_name=sheet_name)

            store_data = pd.read_excel(store_file_path, header=0)
            col = store_data.columns
            collect_data_today = self.set_col(self.format_date_today(collect_data))
            store_data = self.format_date_today(store_data)
            updated_data = _(store_data, collect_data_today, col)
            yield pd.concat([collect_data, updated_data], ignore_index=True), sheet_name


if __name__ == '__main__':
    """
    
    
    
    """
    sye = syncExcel(r'../data/sync_excel', 'file0.xlsx')
    collect_file = ('file0.xlsx', r"../data/sync_excel/file0.xlsx")


    def check_time_to_next_zero():
        now = datetime.datetime.now()
        next_zero = datetime.datetime(now.year, now.month, now.day + 1)
        time_diff = (next_zero - now).total_seconds()
        return time_diff


    while 1:
        count_down = check_time_to_next_zero()
        if count_down <= 31:
            sye.create_not_exists_sheet_on_collect_file(collect_file, list(sye.sheet_dict.keys()))
            result = sye.get_updated_data_body(collect_file)
            with pd.ExcelWriter(collect_file[1], engine='openpyxl', mode='a',
                                if_sheet_exists='replace') as w:
                for data, sheet in result:
                    data.to_excel(w, sheet_name=sheet, index=False, )
        else:
            print(f'当前执行用户: **-->>  {sye.user}  <<--**')
            print(count_down)
            time.sleep(30)
